import os
import logging
from  openpyxl import load_workbook,workbook
from openpyxl.styles import Border,Side
from openpyxl.utils import get_column_letter
import unicodedata
from pathlib import Path



__author__= "evanlinpek"

def get_xls_file_in_current_path():
    '''
    get xls file in current path.
    :return: file path.
    '''
    current_path = os.getcwd()
    all_items = os.listdir(current_path)
    files = [f for f in all_items if f.endswith('.xls') and os.path.isfile(os.path.join(current_path,f))]
    logging.info("all xls files in current path: {}".format(files))
    if len(files) >1:
        raise Exception("Exist multiple xls files,please check,and reserve a correct file!")
    return files.pop()

def add_cess_border_to_xlsx(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    thin_border = Border( left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    wb.save(file_path)
    logging.info('excel file:{},added border,done!')
    return file_path

def unicode_len(s):
    '''
    count chinese char
    :param s:
    :return:
    '''
    return len(unicodedata.normalize('NFC',s))

def adjust_columns_width(file_path):
    '''
    adjust the column width.
    :param file_path: Excel file path
    :return: file_path.
    '''
    wb = load_workbook(file_path)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                cell_length = unicode_len(str(cell.value))
                if cell_length > max_length:max_length = cell_length
            except:
                raise Exception("blank in cell!")
        adjusted_width = (max_length +2) * 1.5
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width
    wb.save(file_path)
    logging.info('adjusted the width of column!')
    return file_path