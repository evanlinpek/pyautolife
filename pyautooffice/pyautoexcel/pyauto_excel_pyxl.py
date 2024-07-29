import os
import logging
from pprint import pprint

from  openpyxl import load_workbook,workbook
from openpyxl.styles import Border,Side
from openpyxl.utils import get_column_letter
import unicodedata
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

__author__= "evanlinpek"
"""
"""

class ExcelOperator:

    def __init__(self,filename = None):
        self.filename = filename
        self.wb = None
        if filename:
            self._load_workbook(filename)

    def create_workbook(self):
        '''
        Create a new workbook.
        :return: None
        '''
        self.wb = Workbook()

    def add_sheet(self, sheet_name,index=None):
        '''
        Create a worksheet (at an optional index)
        :param sheet_name: sheet name
        :param index: optional position at which the sheet will be inserted.
         default,insert at the end;0,insert at first position; -1,insert at the penultimate position
         type index: int
        :return: None
        '''
        if not self.wb:
            self.create_workbook()
        self.wb.create_sheet(sheet_name,index)

    def _load_workbook(self, filename):
        '''
        Load a exist excel workbook
        :param filename: excel filename.
        :return: None
        '''
        self.filename = filename
        self.wb = load_workbook(filename)

    def read_cell(self, sheet_name, row, column):
        """read cell data by row,column"""
        if not self.wb:
            raise ValueError("No workbook is loaded or created.")
        sheet = self.wb[sheet_name]
        return sheet.cell(row=row, column=column).value

    def write_tree_data(self,sheet_name,tree_data):
        if not tree_data:
            raise ValueError("No data")
        if not isinstance(tree_data,list):
            max_depth, depth_info = analyze_list_structure(tree_data)
            if max_depth !=2 or depth_info.get(2):
                raise ValueError("No correct list data,that is not a list of lists")
        if not self.wb:
            raise ValueError("No workbook is loaded or created.")
        sheet = self.wb[sheet_name]
        for row in tree_data:



    def write_cell(self, sheet_name, row, column, value, style=None):
        """write data into cell and can set style."""
        if not self.wb:
            raise ValueError("No workbook is loaded or created.")
        sheet = self.wb[sheet_name]
        cell = sheet.cell(row=row, column=column)
        cell.value = value
        if style:
            cell.font = style['font']
            cell.border = style['border']
            cell.fill = style['fill']
            cell.alignment = style['alignment']

    def set_cell_style(self, sheet_name, row, column, font=None, border=None, fill=None, alignment=None):
        """设置单元格样式"""
        sheet = self.wb[sheet_name]
        cell = sheet.cell(row=row, column=column)
        if font:
            cell.font = Font(**font)
        if border:
            cell.border = Border(left=Side(**border), right=Side(**border), top=Side(**border), bottom=Side(**border))
        if fill:
            cell.fill = PatternFill(**fill)
        if alignment:
            cell.alignment = Alignment(**alignment)


    def save_workbook(self, filename=None):
        """save workbook to a file"""
        if not self.wb:
            raise ValueError("No workbook is loaded or created.")
        if not filename:
            filename = self.filename
        self.wb.save(filename)

def get_xls_file_in_current_path():
    '''
    get xls file in current path.
    :return: file path.
    '''
    current_path = os.getcwd()
    all_items = os.listdir(current_path)
    files = [f for f in all_items if f.endswith('.xls') and os.path.isfile(os.path.join(current_path,f))]
    logging.info("all xls files in current path: {}".format(files))
    if len(files) >1:
        raise Exception("Exist multiple xls files,please check,and reserve a correct file!")
    return files.pop()

def add_cell_border_to_xlsx(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    thin_border = Border( left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
    wb.save(file_path)
    logging.info('excel file:{},added border,done!')
    return file_path

def unicode_len(s):
    '''
    count chinese char
    :param s:
    :return:
    '''
    return len(unicodedata.normalize('NFC',s))

def adjust_columns_width(file_path):
    '''
    adjust the column width.
    :param file_path: Excel file path
    :return: file_path.
    '''
    wb = load_workbook(file_path)
    ws = wb.active
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                cell_length = unicode_len(str(cell.value))
                if cell_length > max_length:max_length = cell_length
            except:
                raise Exception("blank in cell!")
        adjusted_width = (max_length +2) * 1.5
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width
    wb.save(file_path)
    logging.info('adjusted the width of column!')
    return file_path

def analyze_list_structure(lst):
    '''
    analyze list structure
    :param lst: list data.
    :return:
    '''
    def _analyze(current_list, depth):
        if not isinstance(current_list, list):
            return depth - 1  # 非列表元素，返回上一层的深度

        max_depth = depth
        list_count = 0

        for item in current_list:
            current_depth = _analyze(item, depth + 1)
            max_depth = max(max_depth, current_depth)
            if isinstance(item, list):
                list_count += 1

        depth_info[depth] = depth_info.get(depth, 0) + list_count
        return max_depth

    depth_info = {}
    max_depth = _analyze(lst, 1)
    return max_depth, depth_info





if __name__ == "__main__":

    # 创建一个新的Excel操作对象
    eo = ExcelOperator()

    # 创建一个新的工作簿
    eo.create_workbook()

    # 添加一个新的工作表
    eo.add_sheet("MySheet")

    # 设置单元格样式
    cell_style = {
        'font': {'bold': True, 'size': 14},
        'border': {'border_style': 'thin', 'color': '000000'},
        'fill': {'fill_type': 'solid', 'start_color': 'FFFF00'},
        'alignment': {'horizontal': 'center', 'vertical': 'center'}
    }

    # 写入数据到单元格并应用样式
    eo.write_cell("MySheet", 1, 1, "Hello", cell_style)
    eo.write_cell("MySheet", 1, 2, "World", cell_style)

    # 保存工作簿
    eo.save_workbook("formatted_example.xlsx")

